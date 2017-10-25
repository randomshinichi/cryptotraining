import requests
from openpyxl import Workbook

cmc_url = "https://api.coinmarketcap.com/v1/ticker/"
resp = requests.get(cmc_url)
coins = resp.json()

wb = Workbook()
wb.guess_types = True
ws = wb.active

order = ['id', 'name', 'symbol', 'rank', 'price_usd', 'price_btc', '24h_volume_usd', 'market_cap_usd',
         'available_supply', 'total_supply', 'percent_change_1h', 'percent_change_24h', 'percent_change_7d']

# todo: custom cell functions depending on the column.

column = 1
for o in order:
    ws.cell(row=1, column=column, value=o)
    column += 1

for idx, coin in enumerate(coins):
    # import ipdb; ipdb.set_trace()
    column = 1
    for o in order:
        ws.cell(row=idx+2, column=column, value=coin[o])
        column += 1

wb.save('coinmarketcap.xlsx')
